import sys
import os
import pytest
import openpyxl
from unittest.mock import MagicMock

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))
sys.modules.setdefault("manual", MagicMock())

from cp import verifyFile, verifyFiles, getFileList, closeWorkbook, copyImages

TESTFOLDER = os.path.join(os.path.dirname(__file__), "..", "testfolder")
JPG = os.path.join(TESTFOLDER, "wlp.jpg")
PNG = os.path.join(TESTFOLDER, "meme.png")
XLSX = os.path.join(TESTFOLDER, "test_demo.xlsx")


# ---------------------------------------------------------------------------
# verifyFile
# ---------------------------------------------------------------------------

class TestVerifyFile:
    def test_valid_jpg(self):
        assert verifyFile(JPG) == 0

    def test_valid_png(self):
        assert verifyFile(PNG) == 0

    def test_valid_jpeg_extension(self, tmp_path):
        f = tmp_path / "img.jpeg"
        f.write_bytes(b"\xff\xd8\xff")
        assert verifyFile(str(f)) == 0

    def test_missing_file(self, tmp_path):
        assert verifyFile(str(tmp_path / "ghost.jpg")) == 1

    def test_wrong_extension(self, tmp_path):
        f = tmp_path / "data.txt"
        f.write_text("hello")
        assert verifyFile(str(f)) == 2

    def test_xlsx_is_invalid(self):
        assert verifyFile(XLSX) == 2

    def test_case_insensitive_extension(self, tmp_path):
        f = tmp_path / "img.JPG"
        f.write_bytes(b"\xff\xd8\xff")
        assert verifyFile(str(f)) == 0


# ---------------------------------------------------------------------------
# verifyFiles
# ---------------------------------------------------------------------------

class TestVerifyFiles:
    def test_empty_list(self):
        assert verifyFiles([]) == 0

    def test_all_valid(self):
        assert verifyFiles([JPG, PNG]) == 0

    def test_one_missing(self, tmp_path):
        result = verifyFiles([JPG, str(tmp_path / "missing.jpg")])
        assert result == 1

    def test_one_bad_extension(self, tmp_path):
        bad = tmp_path / "doc.pdf"
        bad.write_bytes(b"%PDF")
        assert verifyFiles([JPG, str(bad)]) == 2

    def test_stops_at_first_invalid(self, tmp_path):
        bad = tmp_path / "bad.txt"
        bad.write_text("x")
        result = verifyFiles([str(bad), JPG])
        assert result != 0


# ---------------------------------------------------------------------------
# getFileList
# ---------------------------------------------------------------------------

class TestGetFileList:
    def test_nonexistent_path(self, tmp_path):
        assert getFileList(str(tmp_path / "nope")) == []

    def test_path_is_file(self):
        assert getFileList(JPG) == []

    def test_empty_directory(self, tmp_path):
        assert getFileList(str(tmp_path)) == []

    def test_returns_images_only(self, tmp_path):
        (tmp_path / "a.jpg").write_bytes(b"\xff\xd8\xff")
        (tmp_path / "b.png").write_bytes(b"\x89PNG")
        (tmp_path / "c.txt").write_text("ignored")
        result = getFileList(str(tmp_path))
        assert len(result) == 2
        assert all(f.endswith((".jpg", ".png")) for f in result)

    def test_results_are_sorted(self, tmp_path):
        for name in ("z.jpg", "a.jpg", "m.png"):
            (tmp_path / name).write_bytes(b"\xff\xd8\xff")
        result = getFileList(str(tmp_path))
        assert result == sorted(result)

    def test_full_paths_returned(self, tmp_path):
        img = tmp_path / "pic.jpg"
        img.write_bytes(b"\xff\xd8\xff")
        result = getFileList(str(tmp_path))
        assert result == [str(img)]

    def test_testfolder_returns_images(self):
        result = getFileList(TESTFOLDER)
        assert len(result) >= 3
        assert all(f.lower().endswith((".jpg", ".jpeg", ".png")) for f in result)


# ---------------------------------------------------------------------------
# closeWorkbook
# ---------------------------------------------------------------------------

class TestCloseWorkbook:
    def test_none_workbook_does_not_raise(self, tmp_path):
        closeWorkbook(None, str(tmp_path / "out.xlsx"))

    def test_saves_workbook(self, tmp_path):
        out = tmp_path / "out.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "Sheet1"
        closeWorkbook(wb, str(out))
        assert out.exists()

    def test_saved_file_is_valid_xlsx(self, tmp_path):
        out = tmp_path / "out.xlsx"
        wb = openpyxl.Workbook()
        wb.active["A1"] = "hello"
        closeWorkbook(wb, str(out))
        loaded = openpyxl.load_workbook(str(out))
        assert loaded.active["A1"].value == "hello"


# ---------------------------------------------------------------------------
# copyImages (Auto mode)
# ---------------------------------------------------------------------------

class TestCopyImagesAuto:
    @pytest.fixture()
    def dest(self, tmp_path):
        dst = tmp_path / "dest.xlsx"
        wb = openpyxl.Workbook()
        wb.save(str(dst))
        return str(dst)

    def test_returns_zero_on_success(self, dest):
        assert copyImages([JPG], dest, mode="Auto") == 0

    def test_returns_error_on_bad_xlsx(self, tmp_path):
        bad = tmp_path / "not_an_excel.xlsx"
        bad.write_text("garbage")
        assert copyImages([JPG], str(bad), mode="Auto") == 1

    def test_creates_photos_sheet_if_missing(self, dest):
        wb_before = openpyxl.load_workbook(dest)
        if "Photos" in wb_before.sheetnames:
            del wb_before["Photos"]
            wb_before.save(dest)
        copyImages([JPG], dest, mode="Auto")
        wb_after = openpyxl.load_workbook(dest)
        assert "Photos" in wb_after.sheetnames

    def test_reuses_existing_photos_sheet(self, dest):
        wb = openpyxl.load_workbook(dest)
        if "Photos" not in wb.sheetnames:
            wb.create_sheet("Photos")
            wb.save(dest)
        copyImages([JPG], dest, mode="Auto")
        wb_after = openpyxl.load_workbook(dest)
        assert wb_after.sheetnames.count("Photos") == 1

    def test_multiple_images_stacked_vertically(self, dest):
        assert copyImages([JPG, PNG], dest, colNum=1, rowNum=1, mode="Auto") == 0
        wb = openpyxl.load_workbook(dest)
        ws = wb["Photos"]
        assert len(ws._images) == 2

    def test_custom_start_row_and_col(self, dest):
        assert copyImages([JPG], dest, rowNum=3, colNum=2, mode="Auto") == 0
        wb = openpyxl.load_workbook(dest)
        ws = wb["Photos"]
        img = ws._images[0]
        # openpyxl stores anchors as 0-based; cell(row=3, col=2) → "B3" → row=2, col=1
        assert img.anchor._from.row == 2
        assert img.anchor._from.col == 1
