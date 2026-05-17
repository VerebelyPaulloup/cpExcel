import openpyxl
from openpyxl import load_workbook
import os 
from manual import ManualPlacementWindow

def getFileList(folderPath):
    if not os.path.exists(folderPath):
        print(f"Folder not found: {folderPath}")
        return []
    if not os.path.isdir(folderPath):
        print(f"Not a folder: {folderPath}")
        return []
    
    fileList = []
    for file in sorted(os.listdir(folderPath)):
        fullPath = os.path.join(folderPath, file)
        code = verifyFile(fullPath)
        if code == 0:
            fileList.append(fullPath)
        else:
            print(f"Skipping invalid file: {file}, in folder: {folderPath}, error code: {code}")
    return fileList

def verifyFile(file:str):
    if not os.path.exists(file):
        print(f"File not found: {file}")
        return 1
    if not file.lower().endswith(('.png', '.jpg', '.jpeg')):
        print(f"Invalid image file: {file}")
        return 2
    print(f"File verified: {file}")
    return 0

def verifyFiles(filelist):
    for file in filelist:
        code = verifyFile(file)
        if code != 0:
            return code
        print(f"File verified: {file}")
    return 0

def closeWorkbook(workbook,destinationFile):
    if workbook is None:
        print("No workbook to save.")
        return
    try:
        workbook.save(destinationFile)
        print(f"Workbook saved successfully: {destinationFile}")
    except Exception as e:
        print(f"Error saving workbook: {e}")

def copyImages(imageList, destinationFile, sheetName=None, rowNum=1, colNum=1, mode="Auto",parent=None):
    try:
        wb = load_workbook(destinationFile)
    except Exception as e:
        print(f"Error opening workbook: {e}")
        return 1

    if mode == "Auto":
        ws = wb["Photos"] if "Photos" in wb.sheetnames else wb.create_sheet("Photos")
        for idx, image in enumerate(imageList):
            img = openpyxl.drawing.image.Image(image)
            cell = ws.cell(row=rowNum + idx, column=colNum)
            ws.add_image(img, cell.coordinate)
            ws.row_dimensions[cell.row].height = img.height * 0.75

    elif mode == "Manuel":
        mWindow = ManualPlacementWindow(parent)
        mWindow.showPlacelements(imageList,sheetlist=wb.sheetnames)
        parent.wait_window(mWindow)
        result = mWindow.result
        for item in result:
            img = openpyxl.drawing.image.Image(item["file"])
            ws = wb[item["sheet"]]
            cell = ws.cell(row=int(item["row"]), column=int(item["col"]))
            ws.add_image(img, cell.coordinate)
            ws.row_dimensions[cell.row].height = img.height * 0.75

    closeWorkbook(wb, destinationFile)
    return 0